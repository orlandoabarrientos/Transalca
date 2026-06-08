from flask import Blueprint, request, jsonify, session, Response
from model.report_model import ReportModel

from datetime import datetime
import csv
import io

report_bp = Blueprint('reports', __name__)
model = ReportModel()



@report_bp.route('/dashboard', methods=['GET'])
def dashboard():
    try:
        return jsonify({"status": "success", "data": model.get_dashboard_stats()})
    except Exception as e:
        return jsonify({"status": "error", "message": "No se pudo completar la solicitud."}), 500


@report_bp.route('/recent-orders', methods=['GET'])
def recent_orders():
    try:
        limit = int(request.args.get('limit', 10))
        return jsonify({"status": "success", "data": model.get_recent_orders(limit)})
    except Exception as e:
        return jsonify({"status": "error", "message": "No se pudo completar la solicitud."}), 500


@report_bp.route('/query', methods=['GET'])
def query_reports():
    try:
        if 'user_id' not in session:
            return jsonify({"status": "error", "message": "No autorizado"}), 401

        report_type = request.args.get('type')
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        status = request.args.get('status')

        data = []
        if report_type == 'sales':
            data = model.get_sales_history(start_date, end_date, status)
        elif report_type == 'payments':
            data = model.get_payments_history(start_date, end_date, status)
        elif report_type == 'inventory':
            data = model.get_inventory_kardex(start_date, end_date)
        elif report_type == 'mechanics':
            data = model.get_mechanics_performance(start_date, end_date)
        elif report_type == 'bitacora':
            data = model.get_bitacora_audit(start_date, end_date, status)
        else:
            return jsonify({"status": "error", "message": "Tipo de reporte invalido"}), 400

        return jsonify({"status": "success", "data": data})
    except Exception as e:
        return jsonify({"status": "error", "message": "No se pudo completar la solicitud."}), 500


@report_bp.route('/export', methods=['GET'])
def export_reports():
    try:
        if 'user_id' not in session:
            return jsonify({"status": "error", "message": "No autorizado"}), 401

        report_type = request.args.get('type')
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        status = request.args.get('status')
        format_type = request.args.get('format', 'csv')


        data = []
        filename = f"reporte_{report_type}_{datetime.now().strftime('%Y%m%d%H%M')}"
        headers = []

        if report_type == 'sales':
            data = model.get_sales_history(start_date, end_date, status)
            headers = ["ID", "Cliente", "Fecha", "Total", "Estado"]
            rows = [[d['id'], d['cliente'], d['fecha'], d['total'], d['estado']] for d in data]
            title = "Reporte Orden de Venta"
        elif report_type == 'payments':
            data = model.get_payments_history(start_date, end_date, status)
            headers = ["ID", "Orden", "Cliente", "Fecha", "Referencia", "Monto", "Moneda", "Método", "Estado"]
            rows = [[d['id'], d['orden_id'], d['cliente'], d['fecha'], d['referencia'], d['monto'], d['moneda'], d['metodo'], d['estado']] for d in data]
            title = "Flujo de Pagos"
        elif report_type == 'inventory':
            data = model.get_inventory_kardex(start_date, end_date)
            headers = ["ID", "Producto", "Código", "Motivo", "Tipo Obra", "Fecha", "Cantidad"]
            rows = [[d['id'], d['producto'], d['codigo'], d['motivo'], d['tipo'], d['fecha'], d['cantidad']] for d in data]
            title = "Kardex de Stock"
        elif report_type == 'mechanics':
            data = model.get_mechanics_performance(start_date, end_date)
            headers = ["Mecánico", "Servicios Asignados", "Completados", "Ingreso Generado"]
            rows = [[d['mecanico_nombre'], d['total_asignados'], d['total_completados'], d['ingreso_generado']] for d in data]
            title = "Desempeño de Mecánicos"
        elif report_type == 'bitacora':
            data = model.get_bitacora_audit(start_date, end_date, status)
            headers = ["ID", "Fecha", "Usuario", "Módulo", "Acción", "Descripción", "IP"]
            rows = [[d['id'], d['fecha'], d['usuario'], d['modulo'], d['accion'], d['descripcion'], d['ip']] for d in data]
            title = "Auditoría de Bitácora"
        else:
            return jsonify({"status": "error", "message": "Tipo de reporte invalido"}), 400

        if format_type == 'csv':
            output = io.StringIO()
            writer = csv.writer(output, lineterminator='\n')
            writer.writerow(headers)
            writer.writerows(rows)
            return Response(output.getvalue().encode('utf-8'), mimetype="text/csv", headers={"Content-Disposition": f"attachment;filename={filename}.csv"})

        elif format_type == 'excel':
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = title

            ws.append(headers)

            font_family = "Segoe UI"
            header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="1A365D", end_color="1A365D", fill_type="solid")
            data_font = Font(name=font_family, size=10)
            
            thin_border = Border(
                left=Side(style='thin', color='E2E8F0'),
                right=Side(style='thin', color='E2E8F0'),
                top=Side(style='thin', color='E2E8F0'),
                bottom=Side(style='thin', color='E2E8F0')
            )

            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

            for r in rows:
                ws.append(r)

            currency_cols = []
            date_cols = []
            center_cols = []
            
            for col_idx, header in enumerate(headers, 1):
                header_lower = header.lower()
                if "total" in header_lower or "monto" in header_lower or "ingreso" in header_lower:
                    currency_cols.append(col_idx)
                elif "fecha" in header_lower:
                    date_cols.append(col_idx)
                elif header_lower in ["id", "orden", "código", "estado", "moneda", "ip", "cantidad", "servicios asignados", "completados"]:
                    center_cols.append(col_idx)

            for row_idx in range(2, ws.max_row + 1):
                for col_idx in range(1, len(headers) + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.font = data_font
                    cell.border = thin_border
                    
                    if col_idx in currency_cols:
                        cell.alignment = Alignment(horizontal="right", vertical="center")
                        val = cell.value
                        if isinstance(val, str):
                            val = val.replace('$', '').replace(',', '').strip()
                        try:
                            cell.value = float(val)
                            cell.number_format = '$#,##0.00'
                        except (ValueError, TypeError):
                            pass
                    elif col_idx in date_cols:
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                        val = cell.value
                        if isinstance(val, str):
                            cell.value = val.replace('T', ' ')
                    elif col_idx in center_cols:
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                    else:
                        cell.alignment = Alignment(horizontal="left", vertical="center")

            for col in ws.columns:
                max_len = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    if cell.value is not None:
                        val_str = str(cell.value)
                        if cell.column in currency_cols:
                            val_str = f"${val_str}"
                        max_len = max(max_len, len(val_str))
                ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

            ws.row_dimensions[1].height = 25
            for r_idx in range(2, ws.max_row + 1):
                ws.row_dimensions[r_idx].height = 20

            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            return Response(output.read(), mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f"attachment;filename={filename}.xlsx"})

        elif format_type == 'pdf':
            from fpdf import FPDF
            pdf = FPDF(orientation='L' if report_type in ['bitacora', 'payments', 'inventory'] else 'P')
            pdf.add_page()
            pdf.set_font("helvetica", "B", 16)
            pdf.cell(0, 10, title, new_x="LMARGIN", new_y="NEXT", align="C")
            pdf.set_font("helvetica", "", 10)
            pdf.cell(0, 10, f"Generado: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", new_x="LMARGIN", new_y="NEXT", align="C")
            pdf.ln(5)

            page_w = 277 if pdf.cur_orientation == 'L' else 190
            col_w = page_w / len(headers)

            pdf.set_font("helvetica", "B", 9)
            pdf.set_fill_color(26, 54, 93)
            pdf.set_text_color(255, 255, 255)
            for h in headers:
                pdf.cell(col_w, 8, str(h), border=1, align='C', fill=True)
            pdf.ln(8)

            pdf.set_font("helvetica", "", 8)
            pdf.set_text_color(0, 0, 0)
            for r in rows:
                for col_idx, c in enumerate(r):
                    val = str(c)
                    header_lower = headers[col_idx].lower()
                    if ("total" in header_lower or "monto" in header_lower or "ingreso" in header_lower) and isinstance(c, (int, float)):
                        val = f"${c:.2f}"
                    val = val[:(int(col_w/1.5))]
                    pdf.cell(col_w, 6, val, border=1, align='R' if "total" in header_lower or "monto" in header_lower or "ingreso" in header_lower else 'L')
                pdf.ln(6)

            output = bytearray(pdf.output())
            return Response(bytes(output), mimetype="application/pdf", headers={"Content-Disposition": f"attachment;filename={filename}.pdf"})

        return jsonify({"status": "error", "message": "Formato de exportacion invalido"}), 400

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"status": "error", "message": "No se pudo completar la solicitud."}), 500
